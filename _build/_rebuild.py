import openpyxl, json, os, io, sys, csv, urllib.request, time
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = r"g:\Shared drives\IAMSYbElectronics\Projects\LPFilters_Elliptical_Mike"
SRC  = os.path.join(BASE, "Parts & Simulation Spreadsheet.xlsx")
BOARDS = ["1kHz","2kHz","5kHz","10kHz","20kHz","50kHz","100kHz","200kHz","500kHz","1MHz"]
REFS = ["L1","L2","C1","C2","C3","C4","C5"]

# ---------- re-extract source data ----------
wbs = openpyxl.load_workbook(SRC, data_only=True)
def parse_section(ws, start, end):
    out = {}
    for r in range(start, end+1):
        a = ws.cell(r,1).value
        if not a or "/" not in str(a):
            continue
        ref,_,val = str(a).partition("/")
        out[ref.strip()] = {"value": val.strip(), "dcr_esr": ws.cell(r,2).value,
                            "footprint": ws.cell(r,3).value, "part": ws.cell(r,4).value}
    return out
data = {}
allparts=set()
for b in BOARDS:
    ws=wbs[b]; other_hdr=None
    for r in range(1,30):
        if str(ws.cell(r,1).value).strip().lower().startswith("other choices"):
            other_hdr=r; break
    main=parse_section(ws,2,(other_hdr-1) if other_hdr else 8)
    other=parse_section(ws,(other_hdr+1) if other_hdr else 99,30) if other_hdr else {}
    data[b]={"main":main,"other":other}
    for sec in (main,other):
        for d in sec.values():
            if d["part"]: allparts.add(str(d["part"]).strip())
bws=wbs["Backup Parts"]
for r in range(3,40):
    if bws.cell(r,4).value: allparts.add(str(bws.cell(r,4).value).strip())

# ---------- LCSC info (cached) ----------
cache_path=os.path.join(BASE,"_lcsc_cache.json")
lcsc=json.load(open(cache_path,encoding="utf-8")) if os.path.exists(cache_path) else {}
hdr={"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64)","Accept":"application/json"}
def fetch(code):
    req=urllib.request.Request(f"https://wmsc.lcsc.com/ftps/wm/product/detail?productCode={code}",headers=hdr)
    r=json.loads(urllib.request.urlopen(req,timeout=25).read()).get("result")
    if not r: return {"_error":"null"}
    return {"productModel":r.get("productModel"),"brand":r.get("brandNameEn"),"type":r.get("catalogName"),
            "parentType":r.get("parentCatalogName"),"package":r.get("encapStandard"),
            "description":r.get("productNameEn"),"stock":r.get("stockNumber")}
for p in sorted(allparts):
    if p in lcsc and "_error" not in lcsc[p]: continue
    try: lcsc[p]=fetch(p); time.sleep(0.3)
    except Exception: lcsc[p]={"_error":"fetch failed"}
json.dump(lcsc,open(cache_path,"w",encoding="utf-8"),ensure_ascii=False,indent=1)

FALLBACK={"C2045615":{"type":"Fixed Inductors","parentType":"Inductors","stock":"N/A (delisted)",
          "productModel":"","brand":"","package":"","description":"16uH inductor (no longer listed on LCSC)"}}
def info(pn):
    pn=(pn or "").strip(); d=lcsc.get(pn)
    if not d or "_error" in d: return FALLBACK.get(pn,{})
    return d

# ---------- styles ----------
HDR_FILL=PatternFill("solid",fgColor="1F4E78"); SUB_FILL=PatternFill("solid",fgColor="DDEBF7")
HDR_FONT=Font(bold=True,color="FFFFFF",size=10); SUB_FONT=Font(bold=True,color="1F4E78",size=9)
REF_FONT=Font(bold=True,size=10)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
OOS=PatternFill("solid",fgColor="FF8080"); LOWSTOCK=PatternFill("solid",fgColor="FFC7CE")
CENTER=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")
def stock_fill(v):
    if v in (0,"0") or v is None: return OOS
    try:
        if int(v)<100: return LOWSTOCK
    except (ValueError,TypeError): pass
    return None

wb=openpyxl.Workbook()
ATTRS=["Value","Footprint","Part #","Type","Stock"]

# ---------- matrix: REFS in rows (grouped by attr), BOARDS in columns ----------
def build_matrix(ws, section_key, title):
    ws.cell(1,1,title).font=Font(bold=True,size=13,color="1F4E78")
    # header row 3: corner (Ref | Attr) then one column per board
    ws.cell(3,1,"Ref").font=HDR_FONT; ws.cell(3,1).fill=HDR_FILL; ws.cell(3,1).alignment=CENTER; ws.cell(3,1).border=BORDER
    ws.cell(3,2,"Attribute").font=HDR_FONT; ws.cell(3,2).fill=HDR_FILL; ws.cell(3,2).alignment=CENTER; ws.cell(3,2).border=BORDER
    for j,b in enumerate(BOARDS):
        c=ws.cell(3,3+j,b); c.font=HDR_FONT; c.fill=HDR_FILL; c.alignment=CENTER; c.border=BORDER
    r=4
    for ref in REFS:
        r0=r
        for ai,attr in enumerate(ATTRS):
            ac=ws.cell(r,2,attr); ac.font=SUB_FONT; ac.fill=SUB_FILL; ac.alignment=CENTER; ac.border=BORDER
            for j,b in enumerate(BOARDS):
                d=data[b][section_key].get(ref,{})
                pn=(d.get("part") or "").strip(); li=info(pn)
                v=[d.get("value"),d.get("footprint"),pn or None,li.get("type"),li.get("stock")][ai]
                cell=ws.cell(r,3+j,v); cell.border=BORDER; cell.alignment=LEFT if ai==0 else CENTER
                if attr=="Stock":
                    f=stock_fill(v)
                    if f: cell.fill=f
            r+=1
        # merge ref label cell over its 5 attr rows
        ws.merge_cells(start_row=r0,start_column=1,end_row=r-1,end_column=1)
        rc=ws.cell(r0,1,ref); rc.font=REF_FONT; rc.alignment=CENTER; rc.fill=PatternFill("solid",fgColor="F2F2F2")
        for rr in range(r0,r): ws.cell(rr,1).border=BORDER
    ws.column_dimensions["A"].width=6; ws.column_dimensions["B"].width=11
    for j in range(len(BOARDS)):
        ws.column_dimensions[get_column_letter(3+j)].width=15
    ws.freeze_panes="C4"

ws1=wb.active; ws1.title="BOM (Main parts)"
build_matrix(ws1,"main","Low-Pass Filter BOM — Selected (main) parts  |  references in rows, boards in columns")
ws2=wb.create_sheet("BOM (Alt footprints)")
build_matrix(ws2,"other","Alternate footprint choices ('Other choices Possible' in source)")

# ---------- Footprint grid: refs in rows, boards in columns ----------
ws3=wb.create_sheet("Footprint grid")
ws3.cell(1,1,"Footprint per (reference, board) — Main / Alt").font=Font(bold=True,size=13,color="1F4E78")
ws3.cell(3,1,"Ref").font=HDR_FONT; ws3.cell(3,1).fill=HDR_FILL; ws3.cell(3,1).alignment=CENTER; ws3.cell(3,1).border=BORDER
for j,b in enumerate(BOARDS):
    c=ws3.cell(3,2+j,b); c.font=HDR_FONT; c.fill=HDR_FILL; c.alignment=CENTER; c.border=BORDER
r=4
for ref in REFS:
    rc=ws3.cell(r,1,ref); rc.font=REF_FONT; rc.border=BORDER; rc.fill=PatternFill("solid",fgColor="F2F2F2"); rc.alignment=CENTER
    for j,b in enumerate(BOARDS):
        m=data[b]["main"].get(ref,{}); o=data[b]["other"].get(ref,{})
        mf=m.get("footprint") or ""; of=o.get("footprint")
        txt=mf+(f"\n(alt: {of})" if of else "")
        cell=ws3.cell(r,2+j,txt); cell.alignment=CENTER; cell.border=BORDER
    r+=1
ws3.column_dimensions["A"].width=6
for j in range(len(BOARDS)): ws3.column_dimensions[get_column_letter(2+j)].width=17
ws3.freeze_panes="B4"

# ---------- Parts catalog ----------
usage={}
for b in BOARDS:
    for sk,tag in (("main",""),("other"," (alt)")):
        for ref,d in data[b][sk].items():
            pn=(d.get("part") or "").strip()
            if pn: usage.setdefault(pn,[]).append(f"{b}:{ref}{tag}")
ws4=wb.create_sheet("Parts catalog")
ws4.cell(1,1,"Master parts catalog — every JLCPCB/LCSC part used (find/replace lookup)").font=Font(bold=True,size=13,color="1F4E78")
cols=["JLCPCB/LCSC #","Type","Category","MPN (model)","Brand","Package","Description","Stock","Used on (board:ref)"]
for j,h in enumerate(cols):
    c=ws4.cell(3,1+j,h); c.font=HDR_FONT; c.fill=HDR_FILL; c.alignment=CENTER; c.border=BORDER
r=4
for pn in sorted(usage):
    li=info(pn)
    row=[pn,li.get("type"),li.get("parentType"),li.get("productModel"),li.get("brand"),li.get("package"),li.get("description"),li.get("stock"),", ".join(usage[pn])]
    for j,v in enumerate(row):
        cell=ws4.cell(r,1+j,v); cell.border=BORDER; cell.alignment=LEFT if j in (6,8) else CENTER
        if j==7:
            f=stock_fill(v)
            if f: cell.fill=f
    r+=1
for j,w in enumerate([16,18,14,22,12,9,34,11,30]): ws4.column_dimensions[get_column_letter(1+j)].width=w
ws4.freeze_panes="A4"

# ---------- Backup inductors ----------
ws5=wb.create_sheet("Backup inductors")
ws5.cell(1,1,"Backup inductor catalog (from 'Backup Parts') — with live LCSC type & stock").font=Font(bold=True,size=13,color="1F4E78")
cols5=["L Value","DCR (Ω)","Footprint","JLCPCB #","Type","MPN (model)","Brand","Stock"]
for j,h in enumerate(cols5):
    c=ws5.cell(3,1+j,h); c.font=HDR_FONT; c.fill=HDR_FILL; c.alignment=CENTER; c.border=BORDER
r=4
for sr in range(3,40):
    lv=bws.cell(sr,1).value
    if lv is None: continue
    pn=str(bws.cell(sr,4).value).strip() if bws.cell(sr,4).value else ""
    li=info(pn)
    row=[lv,bws.cell(sr,2).value,bws.cell(sr,3).value,pn,li.get("type"),li.get("productModel"),li.get("brand"),li.get("stock")]
    for j,v in enumerate(row):
        cell=ws5.cell(r,1+j,v); cell.border=BORDER; cell.alignment=LEFT if j==5 else CENTER
        if j==7:
            f=stock_fill(v)
            if f: cell.fill=f
    r+=1
for j,w in enumerate([10,9,18,14,18,22,12,11]): ws5.column_dimensions[get_column_letter(1+j)].width=w
ws5.freeze_panes="A4"

# ---------- Summary ----------
def as_int(v):
    try: return int(v)
    except (ValueError,TypeError): return None
oos=[]; low=[]
for pn in sorted(usage):
    s=info(pn).get("stock"); n=as_int(s)
    if n==0 or s in (0,"0"): oos.append(pn)
    elif n is not None and n<100: low.append(pn)
ws6=wb.create_sheet("Summary")
notes=[
 ("Consolidated LPF Parts & Stock",13,True),("",0,False),
 ("Source: 'Parts & Simulation Spreadsheet.xlsx'. Type & Stock fetched live from LCSC (wmsc.lcsc.com) on 2026-06-10.",10,False),("",0,False),
 ("At a glance:",11,True),
 (f"  • {len(BOARDS)} board variants: {', '.join(BOARDS)}",10,False),
 (f"  • {len(REFS)} references per board: {', '.join(REFS)}",10,False),
 (f"  • {len(usage)} unique LCSC part numbers used across all boards (plus backup inductors).",10,False),
 (f"  • Out of stock (0): {len(oos)} part(s){(' — ' + ', '.join(oos)) if oos else ''}",10,False),
 (f"  • Low stock (<100): {len(low)} part(s){(' — ' + ', '.join(low)) if low else ''}",10,False),("",0,False),
 ("Sheets:",11,True),
 ("  • BOM (Main parts): selected part per board. References (L1,L2,C1..C5) in ROWS; each ref has Value / Footprint / Part # / Type / Stock sub-rows. Boards in COLUMNS.",10,False),
 ("  • BOM (Alt footprints): the 'Other choices Possible' alternates (same value, different footprint).",10,False),
 ("  • Footprint grid: compact footprint-only view — fastest way to scan/find-replace footprints across boards.",10,False),
 ("  • Parts catalog: every unique LCSC part #, its type, package, MPN, brand, stock, and where it's used. Use this as the find/replace lookup.",10,False),
 ("  • Backup inductors: the inductor substitution list, with live type & stock added.",10,False),("",0,False),
 ("Stock colour coding: red = 0 in stock (out of stock); pink = under 100 in stock (low).",10,False),("",0,False),
 ("Note: C2045615 (16uH, 500kHz L2) is no longer listed on LCSC — shown as 'N/A (delisted)'.",10,False),
 ("Tip: Use Excel AutoFilter on the 'Parts catalog' to find every place a given footprint/part is used, then replace.",10,False),
]
for i,(t,sz,b) in enumerate(notes,1):
    c=ws6.cell(i,1,t)
    if t: c.font=Font(bold=b,size=sz or 10,color="1F4E78" if b else "000000")
ws6.column_dimensions["A"].width=130

OUT=os.path.join(BASE,"LPF Consolidated Parts & Stock.xlsx")
wb.save(OUT); print("Saved:",OUT)

# tidy CSV (unchanged shape — long format works either orientation)
CSVOUT=os.path.join(BASE,"LPF Parts (long format).csv")
with open(CSVOUT,"w",newline="",encoding="utf-8-sig") as f:
    w=csv.writer(f)
    w.writerow(["Board","Reference","Choice","Value","Footprint","JLCPCB #","Type","Category","MPN","Brand","Package","Description","Stock"])
    for b in BOARDS:
        for sk,tag in (("main","Main"),("other","Alt")):
            for ref in REFS:
                d=data[b][sk].get(ref)
                if not d: continue
                pn=(d.get("part") or "").strip(); li=info(pn)
                w.writerow([b,ref,tag,d.get("value"),d.get("footprint"),pn,li.get("type"),li.get("parentType"),
                            li.get("productModel"),li.get("brand"),li.get("package"),li.get("description"),li.get("stock")])
print("Saved:",CSVOUT)

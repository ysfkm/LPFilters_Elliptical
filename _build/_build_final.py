import openpyxl, json, os, io, sys, csv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE=r"g:\Shared drives\IAMSYbElectronics\Projects\LPFilters_Elliptical_Mike"
BUILD=os.path.dirname(os.path.abspath(__file__))   # caches/scripts live alongside this file (_build/)
SRC =os.path.join(BASE,"Parts & Simulation Spreadsheet.xlsx")
BOARDS=["1kHz","2kHz","5kHz","10kHz","20kHz","50kHz","100kHz","200kHz","500kHz","1MHz"]
REFS=["L1","L2","C1","C2","C3","C4","C5"]

# ---- load caches (from _build/) ----
lcsc=json.load(open(os.path.join(BUILD,"_lcsc_cache.json"),encoding="utf-8"))
jlc =json.load(open(os.path.join(BUILD,"_jlc_cache.json"),encoding="utf-8"))
alts=json.load(open(os.path.join(BUILD,"_alts_results.json"),encoding="utf-8"))
altbom_data=json.load(open(os.path.join(BUILD,"_altbom.json"),encoding="utf-8"))
altbom=altbom_data["altbom"]; altplan=altbom_data["plan"]

FALL_LCSC={"C2045615":{"type":"Fixed Inductors","parentType":"Inductors","stock":"N/A (delisted)",
           "productModel":"","brand":"","package":"","description":"16uH inductor (no longer listed on LCSC)"}}
def L(pn):
    pn=(pn or "").strip(); d=lcsc.get(pn)
    if not d or "_error" in d: return FALL_LCSC.get(pn,{})
    return d
def J(pn):
    return jlc.get((pn or "").strip(),{})

# ---- extract source ----
wbs=openpyxl.load_workbook(SRC,data_only=True)
def parse_section(ws,s,e):
    out={}
    for r in range(s,e+1):
        a=ws.cell(r,1).value
        if not a or "/" not in str(a): continue
        ref,_,val=str(a).partition("/")
        out[ref.strip()]={"value":val.strip(),"dcr_esr":ws.cell(r,2).value,
                          "footprint":ws.cell(r,3).value,"part":ws.cell(r,4).value}
    return out
data={}
for b in BOARDS:
    ws=wbs[b]; oh=None
    for r in range(1,30):
        if str(ws.cell(r,1).value).strip().lower().startswith("other choices"): oh=r;break
    data[b]={"main":parse_section(ws,2,(oh-1) if oh else 8),
             "other":parse_section(ws,(oh+1) if oh else 99,30) if oh else {}}
bws=wbs["Backup Parts"]

import re as _re
def _capval(s):
    m=_re.search(r'([\d.]+)\s*(pf|nf|uf|µf)',str(s).lower())
    if not m: return None
    return float(m.group(1))*{"pf":1e-12,"nf":1e-9,"uf":1e-6,"µf":1e-6}[m.group(2)]
def _bigger(newv, oldv):
    a=_capval(newv); b=_capval(oldv)
    return (a is not None and b is not None and a>b)

# ---- styles ----
HF=PatternFill("solid",fgColor="1F4E78"); SF=PatternFill("solid",fgColor="DDEBF7")
HFONT=Font(bold=True,color="FFFFFF",size=10); SFONT=Font(bold=True,color="1F4E78",size=9)
RFONT=Font(bold=True,size=10)
thin=Side(style="thin",color="BFBFBF"); BD=Border(left=thin,right=thin,top=thin,bottom=thin)
OOS=PatternFill("solid",fgColor="FF8080"); LOW=PatternFill("solid",fgColor="FFC7CE")
BASICFILL=PatternFill("solid",fgColor="C6EFCE"); PREFFILL=PatternFill("solid",fgColor="BDD7EE")
EXTFILL=PatternFill("solid",fgColor="FFF2CC")
C=Alignment(horizontal="center",vertical="center",wrap_text=True); Lf=Alignment(horizontal="left",vertical="center")
def sfill(v):
    if v in (0,"0") or v is None: return OOS
    try:
        if int(v)<100: return LOW
    except (ValueError,TypeError): pass
    return None
def libfill(lib):
    return {"Basic":BASICFILL,"Preferred":PREFFILL,"Extended":EXTFILL}.get(lib)

wb=openpyxl.Workbook()
# attributes now include JLC library type + JLC stock
ATTRS=["Value","Footprint","Part #","Type","JLC Lib","JLC Stock","LCSC Stock"]

def cellvals(d):
    pn=(d.get("part") or "").strip(); li=L(pn); jj=J(pn)
    return [d.get("value"),d.get("footprint"),pn or None,li.get("type"),
            jj.get("lib"),jj.get("jlc_stock"),li.get("stock")]

def build_matrix(ws,key,title):
    ws.cell(1,1,title).font=Font(bold=True,size=13,color="1F4E78")
    ws.cell(3,1,"Ref").font=HFONT; ws.cell(3,1).fill=HF; ws.cell(3,1).alignment=C; ws.cell(3,1).border=BD
    ws.cell(3,2,"Attribute").font=HFONT; ws.cell(3,2).fill=HF; ws.cell(3,2).alignment=C; ws.cell(3,2).border=BD
    for j,b in enumerate(BOARDS):
        c=ws.cell(3,3+j,b); c.font=HFONT; c.fill=HF; c.alignment=C; c.border=BD
    r=4
    for ref in REFS:
        r0=r
        for ai,attr in enumerate(ATTRS):
            ac=ws.cell(r,2,attr); ac.font=SFONT; ac.fill=SF; ac.alignment=C; ac.border=BD
            for j,b in enumerate(BOARDS):
                d=data[b][key].get(ref,{}); vals=cellvals(d)
                v=vals[ai]; cell=ws.cell(r,3+j,v); cell.border=BD
                cell.alignment=Lf if ai==0 else C
                if attr=="JLC Lib":
                    f=libfill(v)
                    if f: cell.fill=f
                elif attr in ("JLC Stock","LCSC Stock"):
                    f=sfill(v)
                    if f: cell.fill=f
            r+=1
        ws.merge_cells(start_row=r0,start_column=1,end_row=r-1,end_column=1)
        rc=ws.cell(r0,1,ref); rc.font=RFONT; rc.alignment=C; rc.fill=PatternFill("solid",fgColor="F2F2F2")
        for rr in range(r0,r): ws.cell(rr,1).border=BD
    ws.column_dimensions["A"].width=6; ws.column_dimensions["B"].width=11
    for j in range(len(BOARDS)): ws.column_dimensions[get_column_letter(3+j)].width=15
    ws.freeze_panes="C4"

def typ_short(t):
    if not t: return None
    return "MLCC" if "Ceramic" in str(t) else ("Inductor" if "Induct" in str(t) else t)

# ---- Combined BOM: per reference an ORIGINAL block stacked above a PROPOSED block ----
wsC=wb.active; wsC.title="BOM (Original vs Proposed)"
wsC.cell(1,1,"BOM — Original (current parts) vs Proposed (Basic-part swap).  Per reference: ORIGINAL block stacked above PROPOSED block.").font=Font(bold=True,size=13,color="1F4E78")
wsC.cell(2,1,"Boards in columns. Original = current design part; Proposed = stock-first Basic pick (caps ±40%, one footprint per ref). Peach = value changed, orange = footprint changed. JLC Lib: Basic=green / Preferred=blue / Extended=yellow. Inductors keep their original part (no Basic option).").font=Font(italic=True,size=9)
ORIG_LBL=PatternFill("solid",fgColor="D9E1F2"); PROP_LBL=PatternFill("solid",fgColor="E2EFDA")
# header (row 4): Ref | Block | Attribute | boards...
for cc,txt in ((1,"Ref"),(2,"Block"),(3,"Attribute")):
    c=wsC.cell(4,cc,txt); c.font=HFONT; c.fill=HF; c.alignment=C; c.border=BD
for j,b in enumerate(BOARDS):
    c=wsC.cell(4,4+j,b); c.font=HFONT; c.fill=HF; c.alignment=C; c.border=BD
ATTRS5=["Value","Footprint","Part #","JLC Lib","JLC Stock"]
r=5
for ref in REFS:
    ref_r0=r
    for block in ("Original","Proposed"):
        blk_r0=r
        for attr in ATTRS5:
            ac=wsC.cell(r,3,attr); ac.font=Font(bold=True,color="1F4E78",size=9); ac.fill=SF; ac.alignment=C; ac.border=BD
            for j,b in enumerate(BOARDS):
                ab=altbom[f"{b}|{ref}"]; p=ab["pick"]
                opn=ab["orig_part"]; oj=J(opn); ol=L(opn); rj=J(p.get("code"))
                v_chg=ab.get("value_changed"); fp_chg=ab.get("fp_changed")
                if block=="Original":
                    lib_orig=oj.get("lib") or "Extended"
                    val={"Value":ab.get("value"),"Footprint":ab.get("orig_fp"),"Part #":opn,
                         "JLC Lib":lib_orig,"JLC Stock":oj.get("jlc_stock")}[attr]
                else:
                    val={"Value":p.get("value"),"Footprint":p.get("footprint"),"Part #":p.get("code"),
                         "JLC Lib":p.get("lib"),"JLC Stock":p.get("stock")}[attr]
                cell=wsC.cell(r,4+j,val); cell.border=BD
                cell.alignment=Lf if attr=="Value" else C
                if attr=="JLC Lib":
                    f=libfill(val)
                    if f: cell.fill=f
                elif attr=="JLC Stock":
                    f=sfill(val)
                    if f: cell.fill=f
                # highlight changes only on the Proposed block
                if block=="Proposed" and attr=="Value" and v_chg:
                    cell.fill=PatternFill("solid",fgColor="FCE4D6"); cell.font=Font(bold=True,color="C55A11")
                if block=="Proposed" and attr=="Footprint" and fp_chg:
                    cell.fill=PatternFill("solid",fgColor="FFD966"); cell.font=Font(bold=True,color="C55A11")
            r+=1
        # block label (column B) spanning its 5 rows
        wsC.merge_cells(start_row=blk_r0,start_column=2,end_row=r-1,end_column=2)
        bc=wsC.cell(blk_r0,2,block); bc.font=Font(bold=True,size=10); bc.alignment=C
        bc.fill=ORIG_LBL if block=="Original" else PROP_LBL
        for rr in range(blk_r0,r): wsC.cell(rr,2).border=BD
    # ref label (column A) spanning both blocks (10 rows)
    wsC.merge_cells(start_row=ref_r0,start_column=1,end_row=r-1,end_column=1)
    rc=wsC.cell(ref_r0,1,ref); rc.font=RFONT; rc.alignment=C; rc.fill=PatternFill("solid",fgColor="F2F2F2")
    for rr in range(ref_r0,r): wsC.cell(rr,1).border=BD
wsC.column_dimensions["A"].width=6; wsC.column_dimensions["B"].width=10; wsC.column_dimensions["C"].width=11
for j in range(len(BOARDS)): wsC.column_dimensions[get_column_letter(4+j)].width=16
wsC.freeze_panes="D5"


# ---- Parts catalog ----
usage={}
for b in BOARDS:
    for sk,tag in (("main",""),("other"," (alt)")):
        for ref,d in data[b][sk].items():
            pn=(d.get("part") or "").strip()
            if pn: usage.setdefault(pn,[]).append(f"{b}:{ref}{tag}")
ws4=wb.create_sheet("Parts catalog")
ws4.cell(1,1,"Master parts catalog — every JLCPCB/LCSC part used (find/replace lookup)").font=Font(bold=True,size=13,color="1F4E78")
cols4=["JLCPCB/LCSC #","JLC Lib","Type","Category","MPN (model)","Brand","Footprint","Description","JLC stock","LCSC stock","Used on (board:ref)"]
for j,h in enumerate(cols4):
    c=ws4.cell(3,1+j,h); c.font=HFONT; c.fill=HF; c.alignment=C; c.border=BD
r=4
for pn in sorted(usage):
    li=L(pn); jj=J(pn)
    row=[pn,jj.get("lib"),li.get("type"),li.get("parentType"),li.get("productModel"),li.get("brand"),
         jj.get("footprint") or li.get("package"),li.get("description"),jj.get("jlc_stock"),li.get("stock"),", ".join(usage[pn])]
    for j,v in enumerate(row):
        cell=ws4.cell(r,1+j,v); cell.border=BD; cell.alignment=Lf if j in (7,10) else C
        if j==1:
            f=libfill(v);  cell.fill=f if f else cell.fill
        if j in (8,9):
            f=sfill(v)
            if f: cell.fill=f
    r+=1
for j,w in enumerate([16,10,18,13,20,12,11,30,11,11,28]): ws4.column_dimensions[get_column_letter(1+j)].width=w
ws4.freeze_panes="A4"

# ---- Backup inductors ----
ws5=wb.create_sheet("Backup inductors")
ws5.cell(1,1,"Backup inductor catalog (from 'Backup Parts') — with live JLCPCB lib type & stock").font=Font(bold=True,size=13,color="1F4E78")
cols5=["L Value","DCR (Ω)","Footprint","JLCPCB #","JLC Lib","Type","MPN","Brand","JLC stock","LCSC stock"]
for j,h in enumerate(cols5):
    c=ws5.cell(3,1+j,h); c.font=HFONT; c.fill=HF; c.alignment=C; c.border=BD
r=4
for sr in range(3,40):
    lv=bws.cell(sr,1).value
    if lv is None: continue
    pn=str(bws.cell(sr,4).value).strip() if bws.cell(sr,4).value else ""
    li=L(pn); jj=J(pn)
    row=[lv,bws.cell(sr,2).value,bws.cell(sr,3).value,pn,jj.get("lib"),li.get("type"),
         li.get("productModel"),li.get("brand"),jj.get("jlc_stock"),li.get("stock")]
    for j,v in enumerate(row):
        cell=ws5.cell(r,1+j,v); cell.border=BD; cell.alignment=Lf if j==6 else C
        if j==4:
            f=libfill(v);  cell.fill=f if f else cell.fill
        if j in (8,9):
            f=sfill(v)
            if f: cell.fill=f
    r+=1
for j,w in enumerate([10,9,18,14,10,18,20,12,11,11]): ws5.column_dimensions[get_column_letter(1+j)].width=w
ws5.freeze_panes="A4"

# ---- Sim results (LTSpice frequency-response comparison + embedded plots) ----
sim_path=os.path.join(BUILD,"_sim_results.json")
FIGDIR=os.path.join(BASE,"Simulation Figures")
if os.path.exists(sim_path):
    sim=json.load(open(sim_path))
    ws8=wb.create_sheet("Sim results")
    ws8.cell(1,1,"LTSpice frequency-response — Ideal (design) vs Actual (current parts) vs Proposed (Basic-part BOM)").font=Font(bold=True,size=13,color="1F4E78")
    ws8.cell(2,1,"−3 dB cutoff & elliptic-notch frequency per config, with % shift of the Proposed BOM vs the Ideal design. Δ≥15% (notch) highlighted — those filters are meaningfully detuned and should be re-checked. Response plots embedded below.").font=Font(italic=True,size=9)
    # header
    cols=["Board",
          "fc Ideal","fc Actual","fc Proposed","Δfc Prop vs Ideal",
          "Notch Ideal","Notch Proposed","ΔNotch Prop vs Ideal",
          "Notch depth Ideal (dB)","Notch depth Proposed (dB)","Flag"]
    for j,h in enumerate(cols):
        c=ws8.cell(4,1+j,h); c.font=HFONT; c.fill=HF; c.alignment=C; c.border=BD
    def pct(new,base):
        if not new or not base: return None
        return round((new-base)/base*100,1)
    def fhz(x):
        if not x: return None
        return f"{x/1e6:.4g} MHz" if x>=1e6 else (f"{x/1e3:.4g} kHz" if x>=1e3 else f"{x:.4g} Hz")
    WARN=PatternFill("solid",fgColor="FFC7CE"); OKF=PatternFill("solid",fgColor="C6EFCE")
    r=5
    for b in BOARDS:
        d=sim.get(b)
        if not d: continue
        i=d["ideal"]; a=d["actual"]; p=d["proposed"]
        dfc=pct(p["f_3dB"],i["f_3dB"]); dno=pct(p["notch_f"],i["notch_f"])
        flag="re-check" if (dno is not None and abs(dno)>=15) else "ok"
        row=[b,fhz(i["f_3dB"]),fhz(a["f_3dB"]),fhz(p["f_3dB"]),f"{dfc}%" if dfc is not None else None,
             fhz(i["notch_f"]),fhz(p["notch_f"]),f"{dno}%" if dno is not None else None,
             i["notch_db"],p["notch_db"],flag]
        for j,v in enumerate(row):
            cell=ws8.cell(r,1+j,v); cell.border=BD; cell.alignment=C
            if j==0: cell.font=Font(bold=True)
            if j==7 and dno is not None and abs(dno)>=15: cell.fill=WARN; cell.font=Font(bold=True,color="9C0006")
            if j==4 and dfc is not None and abs(dfc)>=15: cell.fill=WARN; cell.font=Font(bold=True,color="9C0006")
            if j==10: cell.fill=WARN if flag=="re-check" else OKF; cell.font=Font(bold=True,color="9C0006" if flag=="re-check" else "006100")
        r+=1
    for j,w in enumerate([8,12,12,12,16,12,14,16,18,20,9]): ws8.column_dimensions[get_column_letter(1+j)].width=w
    # embed response plots, two per row below the table
    img_top=r+2
    ws8.cell(img_top-1,1,"Frequency-response plots (Ideal = grey dashed, Actual = blue, Proposed-Basic = red):").font=Font(bold=True,size=11,color="1F4E78")
    rowh=24  # ~ rows per image block
    for idx,b in enumerate(BOARDS):
        png=os.path.join(FIGDIR,f"{b}_0_response.png")
        if not os.path.exists(png): continue
        img=XLImage(png)
        img.width=int(img.width*0.62); img.height=int(img.height*0.62)
        anchor_col = 1 if idx%2==0 else 9   # column A or I
        anchor_row = img_top + (idx//2)*rowh
        img.anchor=f"{get_column_letter(anchor_col)}{anchor_row}"
        ws8.add_image(img)

# ---- Summary ----
def as_int(v):
    try: return int(v)
    except (ValueError,TypeError): return None
basics=[p for p in usage if J(p).get("lib")=="Basic"]
prefs =[p for p in usage if J(p).get("lib")=="Preferred"]
exts  =[p for p in usage if J(p).get("lib")=="Extended"]
oos=[p for p in usage if (as_int(J(p).get("jlc_stock"))==0 or J(p).get("jlc_stock") in (0,"0",None))]
# Alternative BOM stats
ab_basic=sum(1 for v in altbom.values() if v["pick"]["lib"]=="Basic")
ab_pref =sum(1 for v in altbom.values() if v["pick"]["lib"]=="Preferred")
ab_ext  =70-ab_basic-ab_pref
ab_fpchg=sum(1 for v in altbom.values() if v.get("fp_changed"))
ab_vchg =sum(1 for v in altbom.values() if v.get("value_changed"))
ab_vbig =sum(1 for v in altbom.values() if v["pick"].get("value_delta_pct",0)>=20)
ws6=wb.create_sheet("Summary")
notes=[
 ("Consolidated LPF Parts, Library Type & Stock",13,True),("",0,False),
 ("Source: 'Parts & Simulation Spreadsheet.xlsx'. Library type (Basic/Extended/Preferred) & stock from JLCPCB assembly API; extra detail from LCSC. Fetched 2026-06-10.",10,False),("",0,False),
 ("At a glance:",11,True),
 (f"  • {len(BOARDS)} boards x {len(REFS)} references.  {len(usage)} unique part numbers used.",10,False),
 (f"  • JLCPCB library type of parts in use: {len(basics)} Basic, {len(prefs)} Preferred, {len(exts)} Extended.",10,False),
 (f"      Basic parts already used: {', '.join(basics) if basics else 'none'}",10,False),
 (f"  • {len(oos)} part(s) with 0 JLCPCB assembly stock: {', '.join(oos) if oos else 'none'}",10,False),
 ("",0,False),
 ("'BOM (Original vs Proposed)' sheet:",11,True),
 ("  • Per reference: an ORIGINAL block (current design part) stacked above a PROPOSED block (Basic-part swap). Value / Footprint / Part # / JLC Lib / JLC Stock; boards in columns.",10,False),
 (f"  • Proposed = stock-first, Basic-preferred (caps ±40%, one footprint per ref). ALL {ab_basic} capacitor slots become well-stocked Basic (0603) vs only 2 Basic originally; the {ab_ext} remaining Extended are the inductors (no Basic option).",10,False),
 (f"  • ⚠ TRADE-OFF: {ab_vchg} of 50 caps shift >5% from the design value, {ab_vbig} of them ≥20% (e.g. 330nF→220nF, 1.5µF→1µF) — no Basic exists at the exact value. {ab_fpchg} cells also change footprint (→0603). Peach = value changed, orange = footprint changed.",10,False),
 ("  • Inductors keep their per-board footprint (a single size can't span 8.2uH..10mH) and stay as the original Extended parts.",10,False),
 ("",0,False),
 ("'Sim results' sheet: LTSpice frequency-response comparison (Ideal vs Actual vs Proposed-Basic) for all 10 boards — −3dB cutoff & notch shift table plus embedded response plots. Flags filters detuned ≥15% by the Basic swap (1kHz, 5kHz, 20kHz).",10,False),
 ("",0,False),
 ("Colour key:  Basic = green, Preferred = blue, Extended = yellow (JLC Lib cells).  Stock: red = 0, pink = <100.  Footprint change = orange.  Value change = peach.",10,False),
 ("",0,False),
 ("Important: JLCPCB assembly stock != LCSC retail stock. e.g. C14663 reads 0 on LCSC retail but 21.6M in the JLCPCB assembly library. Use the JLC stock column for assembly decisions.",10,False),
 ("Note: C2045615 (16uH, 500kHz L2) is delisted on LCSC retail but still in JLCPCB library (281 pcs).",10,False),
 ("",0,False),
 ("Sheets: BOM (Original vs Proposed) | Sim results | Parts catalog | Backup inductors | Summary.",10,False),
]
for i,(t,sz,bold) in enumerate(notes,1):
    c=ws6.cell(i,1,t)
    if t: c.font=Font(bold=bold,size=sz or 10,color="1F4E78" if bold else "000000")
ws6.column_dimensions["A"].width=140

# order sheets
order=["BOM (Original vs Proposed)","Sim results","Parts catalog","Backup inductors","Summary"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

OUT=os.path.join(BASE,"LPF Consolidated Parts & Stock.xlsx")
try:
    wb.save(OUT)
except PermissionError:
    OUT=os.path.join(BASE,"LPF Consolidated Parts & Stock (v2).xlsx")
    wb.save(OUT)
    print("(original was locked/open — saved a new copy)")
print("Saved:",OUT)
print(f"Sheets: {[s.title for s in wb._sheets]}")

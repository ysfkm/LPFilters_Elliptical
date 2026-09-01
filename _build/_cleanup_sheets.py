import openpyxl, os, sys
sys.stdout.reconfigure(encoding="utf-8")
F=r"g:\Shared drives\IAMSYbElectronics\Projects\LPFilters_Elliptical_Mike\LPF Consolidated Parts & Stock.xlsx"
DROP=["Original vs Replacement","Basic part suggestions","Footprint grid","BOM (Alt footprints)"]
wb=openpyxl.load_workbook(F)
print("Before:",wb.sheetnames)
for name in DROP:
    if name in wb.sheetnames:
        del wb[name]
wb.save(F)
wb2=openpyxl.load_workbook(F)
print("After: ",wb2.sheetnames)
print("Sim results images:",len(wb2["Sim results"]._images))

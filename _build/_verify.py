import openpyxl, io, sys
sys.stdout=io.TextIOWrapper(sys.stdout.buffer,encoding="utf-8")
wb=openpyxl.load_workbook(r"g:\Shared drives\IAMSYbElectronics\Projects\LPFilters_Elliptical_Mike\LPF Consolidated Parts & Stock.xlsx")
ws=wb["Original vs Replacement"]
print(f"=== Original vs Replacement ({ws.max_row}x{ws.max_column}) — L1 + C1 blocks, boards 1kHz..10kHz ===")
for r in range(4,ws.max_row+1):
    ref=ws.cell(r,1).value; blk=ws.cell(r,2).value; at=ws.cell(r,3).value
    cells=[ws.cell(r,c).value for c in range(4,8)]
    print(f"{str(ref or ''):4}{str(blk or ''):10}{str(at or ''):11}| " + " | ".join("" if v is None else str(v) for v in cells))
    if r>=34: break

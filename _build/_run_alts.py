import json, os, io, sys, time
sys.path.insert(0, r"g:\Shared drives\IAMSYbElectronics\Projects\LPFilters_Elliptical_Mike")
import importlib.util
spec=importlib.util.spec_from_file_location("alts", r"g:\Shared drives\IAMSYbElectronics\Projects\LPFilters_Elliptical_Mike\_jlc_alts.py")
alts=importlib.util.module_from_spec(spec); spec.loader.exec_module(alts)
sys.stdout.reconfigure(encoding="utf-8")

BASE=r"g:\Shared drives\IAMSYbElectronics\Projects\LPFilters_Elliptical_Mike"
import openpyxl
SRC=os.path.join(BASE,"Parts & Simulation Spreadsheet.xlsx")
BOARDS=["1kHz","2kHz","5kHz","10kHz","20kHz","50kHz","100kHz","200kHz","500kHz","1MHz"]
REFS=["L1","L2","C1","C2","C3","C4","C5"]

# re-extract main parts (value+footprint+part) from source
wbs=openpyxl.load_workbook(SRC,data_only=True)
def parse_section(ws,s,e):
    out={}
    for r in range(s,e+1):
        a=ws.cell(r,1).value
        if not a or "/" not in str(a): continue
        ref,_,val=str(a).partition("/")
        out[ref.strip()]={"value":val.strip(),"dcr":ws.cell(r,2).value,"footprint":ws.cell(r,3).value,"part":ws.cell(r,4).value}
    return out
main={}
for b in BOARDS:
    ws=wbs[b]; oh=None
    for r in range(1,30):
        if str(ws.cell(r,1).value).strip().lower().startswith("other choices"): oh=r;break
    main[b]=parse_section(ws,2,(oh-1) if oh else 8)

jcache=json.load(open(os.path.join(BASE,"_jlc_cache.json"),encoding="utf-8"))
def lib_of_part(pn):
    d=jcache.get((pn or "").strip(),{})
    return d.get("lib")

# Normalize footprint to the JLCPCB "componentSpecificationEn" form.
# For caps the source footprint is just the size code (0603 etc) embedded in column C like "X7R, 0603".
def cap_fp(footprint_str):
    m=__import__("re").search(r'(0201|0402|0603|0805|1206|1210|1812|2220)',str(footprint_str))
    return m.group(1) if m else None
def is_cap(ref): return ref.startswith("C")

results={}  # key f"{board}|{ref}" -> {"part":..,"lib":..,"alts":[...]}
total=len(BOARDS)*len(REFS); n=0
for b in BOARDS:
    for ref in REFS:
        n+=1
        d=main[b].get(ref,{})
        pn=(d.get("part") or "").strip()
        val=d.get("value"); fp_src=d.get("footprint")
        cur_lib=lib_of_part(pn)
        entry={"board":b,"ref":ref,"part":pn,"value":val,"footprint_src":fp_src,
               "current_lib":cur_lib,"alts":[]}
        # Only search when current part is NOT already Basic/Preferred
        if cur_lib in ("Basic","Preferred"):
            entry["note"]="already Basic/Preferred"
        else:
            if is_cap(ref):
                fp=cap_fp(fp_src)
                entry["footprint"]=fp
                if fp and val:
                    entry["alts"]=alts.find_cap_alts(val,fp)
            else:
                fp=jcache.get(pn,{}).get("footprint")  # inductor footprint from JLC (e.g. SMD,6x6mm)
                entry["footprint"]=fp
                # original DCR (col B): number like 5.85 or string like '88m'
                dcr_raw=d.get("dcr")
                max_dcr=alts.parse_dcr(dcr_raw) if isinstance(dcr_raw,str) else (float(dcr_raw) if isinstance(dcr_raw,(int,float)) else None)
                if fp and val:
                    entry["alts"]=alts.find_ind_alts(val,fp,max_dcr=max_dcr)
        results[f"{b}|{ref}"]=entry
        na=len(entry["alts"])
        print(f"[{n}/{total}] {b} {ref} {val} ({cur_lib}) -> {na} alt(s)" + (f"  best={entry['alts'][0]['code']}({entry['alts'][0]['lib']})" if na else ""))
        time.sleep(0.25)

json.dump(results,open(os.path.join(BASE,"_alts_results.json"),"w",encoding="utf-8"),ensure_ascii=False,indent=1)
print("ALTS DONE")

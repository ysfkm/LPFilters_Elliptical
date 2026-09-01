import json, os, io, sys, time, importlib.util
spec=importlib.util.spec_from_file_location("alts", r"g:\Shared drives\IAMSYbElectronics\Projects\LPFilters_Elliptical_Mike\_jlc_alts.py")
alts=importlib.util.module_from_spec(spec); spec.loader.exec_module(alts)
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

BASE=r"g:\Shared drives\IAMSYbElectronics\Projects\LPFilters_Elliptical_Mike"
SRC=os.path.join(BASE,"Parts & Simulation Spreadsheet.xlsx")
BOARDS=["1kHz","2kHz","5kHz","10kHz","20kHz","50kHz","100kHz","200kHz","500kHz","1MHz"]
REFS=["L1","L2","C1","C2","C3","C4","C5"]

wbs=openpyxl.load_workbook(SRC,data_only=True)
def parse_section(ws,s,e):
    out={}
    for r in range(s,e+1):
        a=ws.cell(r,1).value
        if not a or "/" not in str(a): continue
        ref,_,val=str(a).partition("/")
        out[ref.strip()]={"value":val.strip(),"dcr":ws.cell(r,2).value,"footprint":ws.cell(r,3).value,"part":ws.cell(r,4).value}
    return out
main={}
for b in BOARDS:
    ws=wbs[b]; oh=None
    for r in range(1,30):
        if str(ws.cell(r,1).value).strip().lower().startswith("other choices"): oh=r;break
    main[b]=parse_section(ws,2,(oh-1) if oh else 8)

jcache=json.load(open(os.path.join(BASE,"_jlc_cache.json"),encoding="utf-8"))
def cap_fp(s):
    import re; m=re.search(r'(0201|0402|0603|0805|1206|1210|1812|2220)',str(s)); return m.group(1) if m else None
def is_cap(ref): return ref.startswith("C")
def parse_dcr(v):
    return alts.parse_dcr(v) if isinstance(v,str) else (float(v) if isinstance(v,(int,float)) else None)

# value-keyed cap candidate cache so repeated values (e.g. 330nF appears twice) reuse one query
capcache={}
out={}  # key board|ref -> {value, orig_fp, orig_part, orig_lib, orig_dcr, candidates:[...]}
n=0; total=len(BOARDS)*len(REFS)
for b in BOARDS:
    for ref in REFS:
        n+=1
        d=main[b].get(ref,{})
        pn=(d.get("part") or "").strip()
        val=d.get("value")
        orig_fp = cap_fp(d.get("footprint")) if is_cap(ref) else jcache.get(pn,{}).get("footprint")
        rec={"board":b,"ref":ref,"value":val,"orig_fp":orig_fp,"orig_part":pn,
             "orig_lib":jcache.get(pn,{}).get("lib"),"orig_dcr":parse_dcr(d.get("dcr")),
             "candidates":[]}
        if is_cap(ref) and val:
            if val not in capcache:
                capcache[val]=alts.all_cap_candidates(val)
                time.sleep(0.25)
            rec["candidates"]=capcache[val]
        elif val:  # inductor: gather across footprints via find_ind_alts with no dcr cap & big result set
            cands=alts.find_ind_alts(val, orig_fp or "", max_dcr=None, max_results=40)
            rec["candidates"]=cands
            time.sleep(0.25)
        out[f"{b}|{ref}"]=rec
        nb=sum(1 for c in rec["candidates"] if c["lib"]=="Basic")
        print(f"[{n}/{total}] {b} {ref} {val}: {len(rec['candidates'])} cand ({nb} basic)")

json.dump(out,open(os.path.join(BASE,"_candidates.json"),"w",encoding="utf-8"),ensure_ascii=False,indent=1)
print("GATHER DONE")
